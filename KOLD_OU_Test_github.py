import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------ CONFIG ------------------
# Use relative/sample paths for GitHub/demo; adjust as needed.
input_path = Path("data/sample_mlb_10yrs.xlsx")
output_path = Path("outputs/2025.7.27_KOU_2x_8Ls_ALL_10yrs.xlsx")
input_sheet_name = 0  # Use first sheet
ou_sheet_name = 'O_U'
tail_sheet_name = 'Tails Prior'
results_sheet_name = 'Results'

# ------------------ STEP 1: O/U PIVOT SHEET ------------------
df = pd.read_excel(input_path, sheet_name=input_sheet_name)

# Merge historical “Oakland” / “Sacramento” naming into a single key for continuity
df['Team'] = df['Team'].replace({'Oakland': 'Oakland_Sacramento', 'Sacramento': 'Oakland_Sacramento'})

def over_under_label(margin):
    if margin > 0:
        return "Over"
    elif margin == 0:
        return "Push"
    else:
        return "Under"

df['Over_Under'] = df["O/U Margin"].apply(over_under_label)

pivot_table = df.pivot(index='Date', columns='Team', values='Over_Under')
pivot_table = pivot_table.sort_index().reset_index()

def color_ou(val):
    if val == "Over":
        return 'background-color: lightgreen'
    elif val == "Under":
        return 'background-color: #ffcccc'
    elif val == "Push":
        return 'background-color: #ADD8E6'
    elif val in ("", None):
        return 'background-color: #D3D3D3'
    else:
        return ''

styled = pivot_table.style.applymap(color_ou, subset=pd.IndexSlice[:, pivot_table.columns[1:]])
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    styled.to_excel(writer, sheet_name=ou_sheet_name, index=False)

# ------------------ STEP 2: TAILS PRIOR SYSTEM ------------------
ou_df = pd.read_excel(output_path, sheet_name=ou_sheet_name)
ou_df['Date'] = pd.to_datetime(ou_df['Date'])

teams = [col for col in ou_df.columns if col != 'Date']
dates = ou_df['Date']

def compute_team_tail(series, dates):
    """
    Compute monthly-reset tail instructions per team based on prior actual O/U result.
    - Monthly reset: first non-push game each month is 'Skip' (no prior to tail).
    - If last_result == Over -> 'Tail - Over'
      If last_result == Under -> 'Fade - Under'
    """
    tail_col = []
    last_result = None
    current_month = None
    for idx, (result, date) in enumerate(zip(series, dates)):
        this_month = date.month if not pd.isna(date) else None
        if current_month != this_month:
            current_month = this_month
            last_result = None
        if pd.isna(result) or result == '' or result == 'Skip':
            tail_col.append('Skip')
        elif result == 'Push':
            tail_col.append('Push')
        else:
            if last_result is None:
                tail_col.append('Skip')
            else:
                if last_result == 'Over':
                    tail_col.append('Tail - Over')
                elif last_result == 'Under':
                    tail_col.append('Fade - Under')
                else:
                    tail_col.append('Skip')
            if result in ('Over', 'Under'):
                last_result = result
    return tail_col

# --- Renamed: previously analyze_four_loss_streaks ---
def analyze_eight_loss_recovery(system_series, ou_series, dates):
    """
    Detect sequences where the system records 8 consecutive losses (YEARLY reset),
    then evaluate the next up-to-2 non-push games for at least one win ("recovered").

    Returns:
        flags: list[str] with markers per row ('X' at 8th loss, 'W'/'L' during recovery completion)
        results: list[dict] with {year, start_date, end_date, recovered}
    """
    flags = [''] * len(system_series)
    results = []

    # Track consecutive losses by YEAR (not month)
    current_year = None
    consecutive_losses = 0
    in_recovery = False
    recovery_games = 0
    recovery_wins = 0
    four_loss_start_date = None  # retained var name, now used for 8L start for minimal changes
    recovery_start_idx = None

    for idx, (system_val, ou_val, date) in enumerate(zip(system_series, ou_series, dates)):
        this_year = date.year if not pd.isna(date) else None

        # Reset on year change
        if current_year != this_year:
            current_year = this_year
            consecutive_losses = 0
            in_recovery = False
            recovery_games = 0
            recovery_wins = 0
            four_loss_start_date = None
            recovery_start_idx = None

        # Skip if no valid system prediction or game result
        if system_val in ('Skip', 'Push', '', None) or pd.isna(system_val) or ou_val in ('Skip', '', None) or pd.isna(ou_val):
            continue

        # If currently in recovery mode (2-game window; pushes don't count)
        if in_recovery:
            if ou_val == "Push":
                continue  # does not consume a recovery game
            recovery_games += 1

            # Recovery win conditions (expanded logic)
            is_win = False
            if system_val == "Tail - Over" and ou_val == "Over":
                is_win = True
            elif system_val == "Tail - Under" and ou_val == "Under":
                is_win = True
            elif system_val == "Fade - Under" and ou_val == "Under":
                is_win = True
            elif system_val == "Fade - Over" and ou_val == "Over":
                is_win = True

            if is_win and recovery_wins == 0:  # First win in recovery period
                flags[idx] = 'W'
                recovery_wins += 1
            elif is_win:
                recovery_wins += 1

            # Check if we've completed the recovery period (2 games)
            if recovery_games >= 2:
                recovery_success = recovery_wins > 0

                # If no wins in recovery period, mark the last game as 'L'
                if not recovery_success:
                    flags[idx] = 'L'

                results.append({
                    'year': current_year,
                    'start_date': four_loss_start_date,
                    'end_date': date,
                    'recovered': recovery_success
                })
                in_recovery = False
                recovery_games = 0
                recovery_wins = 0
                four_loss_start_date = None
                recovery_start_idx = None

            continue  # don't double-process

        # Outside recovery: pushes don't count toward streaks
        if ou_val == "Push":
            continue

        # Loss conditions (expanded logic)
        is_loss = False
        if system_val == "Tail - Over" and ou_val == "Under":
            is_loss = True
        elif system_val == "Tail - Under" and ou_val == "Over":
            is_loss = True
        elif system_val == "Fade - Under" and ou_val == "Over":
            is_loss = True
        elif system_val == "Fade - Over" and ou_val == "Under":
            is_loss = True

        if is_loss:
            if consecutive_losses == 0:
                four_loss_start_date = date  # first loss date (now actually 1st of 8L)
            consecutive_losses += 1

            # Mark the 8th consecutive loss and start recovery
            if consecutive_losses == 8:
                flags[idx] = 'X'
                in_recovery = True
                recovery_games = 0
                recovery_wins = 0
                consecutive_losses = 0  # reset for next potential streak
                recovery_start_idx = idx + 1
        else:
            # Any win breaks the current loss streak count
            consecutive_losses = 0

    return flags, results

# Build output, alternating [Team, Team Tail, Team Flag, X]
output_cols = ['Date']
output_data = {'Date': dates}
all_results = []

for team in teams:
    tail_col = compute_team_tail(ou_df[team], dates)
    tail_col_name = f"{team} Tail"
    flag_col_name = f"{team} Flag"
    x_col_name = "X"  # divider retained per your preference

    # Analyze 8-loss streaks for this team (YEARLY reset)
    flags, team_results = analyze_eight_loss_recovery(tail_col, ou_df[team], dates)

    # Add team identifier to results
    for result in team_results:
        result['system'] = 'Tails Prior'
        result['team'] = team
    all_results.extend(team_results)

    output_data[team] = ou_df[team]
    output_data[tail_col_name] = tail_col
    output_data[flag_col_name] = flags
    output_data[x_col_name] = [''] * len(dates)
    output_cols += [team, tail_col_name, flag_col_name, x_col_name]

tail_df = pd.DataFrame(output_data)[output_cols]

with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    tail_df.to_excel(writer, sheet_name=tail_sheet_name, index=False)

# ------------------ STEP 3: COLOR FORMATTING FOR TAILS PRIOR ------------------
def apply_tails_coloring(wb, sheet_name, teams):
    """
    Apply cell fill colors for O/U results and Tail outcomes.
    Green=win, Red=loss, Blue=push, Gray=skip/blank.
    """
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    nrows = ws.max_row

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for team in teams:
        team_col = headers.index(team) + 1
        tail_col = headers.index(f"{team} Tail") + 1

        for row in range(2, nrows+1):
            ou_val = ws.cell(row=row, column=team_col).value
            tail_val = ws.cell(row=row, column=tail_col).value

            # O/U coloring
            if ou_val == "Over":
                ws.cell(row=row, column=team_col).fill = green_fill
            elif ou_val == "Under":
                ws.cell(row=row, column=team_col).fill = red_fill
            elif ou_val == "Push":
                ws.cell(row=row, column=team_col).fill = blue_fill
            elif ou_val in ('Skip', '', None):
                ws.cell(row=row, column=team_col).value = "Skip"
                ws.cell(row=row, column=team_col).fill = gray_fill

            # Tail column coloring by outcome
            if tail_val in ("Tail - Over", "Fade - Under"):
                if tail_val == "Tail - Over":
                    if ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
                elif tail_val == "Fade - Under":
                    if ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Push":
                ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Skip":
                ws.cell(row=row, column=tail_col).fill = gray_fill

wb = load_workbook(output_path)
apply_tails_coloring(wb, tail_sheet_name, teams)
wb.save(output_path)

# ------------------ STEP 4: SYSTEM PATTERN SHEETS ------------------
# Note: Pattern systems reset monthly; recovery analysis resets yearly.
system_patterns = {
    # ORIGINAL PATTERNS
    "OU TFTF": ["Tail - Over", "Fade - Under"],
    "OU TFFT": ["Tail - Over", "Fade - Under", "Fade - Under", "Tail - Over"],
    "OU FTTF": ["Fade - Under", "Tail - Over", "Tail - Over", "Fade - Under"],
}

def generate_system_col(ou_series, dates, pattern):
    """
    Generate monthly-reset pattern instructions irrespective of outcomes.
    """
    result = []
    current_month = None  # monthly reset for patterns
    pattern_idx = 0
    for idx, (ou, date) in enumerate(zip(ou_series, dates)):
        this_month = date.month if not pd.isna(date) else None
        if current_month != this_month:
            current_month = this_month
            pattern_idx = 0
        if ou in ("Skip", '', None) or pd.isna(ou):
            result.append("Skip")
        elif ou == "Push":
            result.append("Push")
        else:
            result.append(pattern[pattern_idx % len(pattern)])
            pattern_idx += 1
    return result

for system_name, pattern in system_patterns.items():
    system_df = tail_df.copy()

    for team in teams:
        tail_col = f"{team} Tail"
        sys_col = f"{team} {system_name}"
        flag_col = f"{team} {system_name} Flag"

        # Generate system column
        sys_series = generate_system_col(system_df[team], system_df['Date'], pattern)

        # Analyze 8-loss streaks for this system/team (YEARLY reset)
        flags, team_results = analyze_eight_loss_recovery(sys_series, system_df[team], system_df['Date'])

        # Add team/system identifier to results
        for result in team_results:
            result['system'] = system_name
            result['team'] = team
        all_results.extend(team_results)

        # Insert system and flag columns next to Tail
        tail_idx = list(system_df.columns).index(tail_col) + 1
        system_df.insert(tail_idx, sys_col, sys_series)
        system_df.insert(tail_idx + 1, flag_col, flags)

    # Save system sheet
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        system_df.to_excel(writer, sheet_name=system_name, index=False)

    # Color formatting for this sheet
    wb = load_workbook(output_path)
    ws = wb[system_name]
    headers = [cell.value for cell in ws[1]]
    nrows = ws.max_row

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for team in teams:
        team_col = headers.index(team) + 1
        tail_col = headers.index(f"{team} Tail") + 1
        sys_col = headers.index(f"{team} {system_name}") + 1

        for row in range(2, nrows+1):
            ou_val = ws.cell(row=row, column=team_col).value
            tail_val = ws.cell(row=row, column=tail_col).value
            sys_val = ws.cell(row=row, column=sys_col).value

            # O/U coloring
            if ou_val == "Over":
                ws.cell(row=row, column=team_col).fill = green_fill
            elif ou_val == "Under":
                ws.cell(row=row, column=team_col).fill = red_fill
            elif ou_val == "Push":
                ws.cell(row=row, column=team_col).fill = blue_fill
            elif ou_val in ('Skip', '', None):
                ws.cell(row=row, column=team_col).value = "Skip"
                ws.cell(row=row, column=team_col).fill = gray_fill

            # Tail column coloring by outcome
            if tail_val in ("Tail - Over", "Fade - Under"):
                if tail_val == "Tail - Over":
                    if ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
                elif tail_val == "Fade - Under":
                    if ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Push":
                ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Skip":
                ws.cell(row=row, column=tail_col).fill = gray_fill

            # System column coloring by O/U result
            if sys_val in ("Tail - Over", "Fade - Under"):
                if sys_val == "Tail - Over":
                    if ou_val == "Over":
                        ws.cell(row=row, column=sys_col).fill = green_fill
                    elif ou_val == "Under":
                        ws.cell(row=row, column=sys_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=sys_col).fill = blue_fill
                elif sys_val == "Fade - Under":
                    if ou_val == "Under":
                        ws.cell(row=row, column=sys_col).fill = green_fill
                    elif ou_val == "Over":
                        ws.cell(row=row, column=sys_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=sys_col).fill = blue_fill
            elif sys_val == "Push":
                ws.cell(row=row, column=sys_col).fill = blue_fill
            elif sys_val == "Skip":
                ws.cell(row=row, column=sys_col).fill = gray_fill

    wb.save(output_path)

# ------------------ STEP 5: PRIOR-BASED SYSTEM PATTERN SHEETS ------------------
# Note: Prior-based uses monthly patterning of Tail/Fade applied to the prior's recommendation.
prior_system_patterns = {
    # ORIGINAL PATTERNS
    "TP TFTF": ["Tail", "Fade"],
    "TP TFFT": ["Tail", "Fade", "Fade", "Tail"],
    "TP FTTF": ["Fade", "Tail", "Tail", "Fade"],
}

def generate_prior_system_col(tail_series, dates, pattern):
    """
    Apply monthly-reset Tail/Fade pattern to the prior-based tail signal.
    """
    result = []
    current_month = None
    pattern_idx = 0

    for idx, (tail_val, date) in enumerate(zip(tail_series, dates)):
        this_month = date.month if not pd.isna(date) else None
        if current_month != this_month:
            current_month = this_month
            pattern_idx = 0

        if tail_val in ("Skip", '', None) or pd.isna(tail_val):
            result.append("Skip")
        elif tail_val == "Push":
            result.append("Push")
        else:
            # Get the pattern action (Tail or Fade)
            pattern_action = pattern[pattern_idx % len(pattern)]

            if tail_val == "Tail - Over":
                if pattern_action == "Tail":
                    result.append("Tail - Over")
                else:  # Fade
                    result.append("Fade - Under")
            elif tail_val == "Fade - Under":
                if pattern_action == "Tail":
                    result.append("Tail - Under")
                else:  # Fade
                    result.append("Fade - Over")
            else:
                result.append("Skip")

            pattern_idx += 1

    return result

for system_name, pattern in prior_system_patterns.items():
    system_df = tail_df.copy()

    for team in teams:
        tail_col = f"{team} Tail"
        sys_col = f"{team} {system_name}"
        flag_col = f"{team} {system_name} Flag"

        # Generate prior-based system column
        sys_series = generate_prior_system_col(system_df[tail_col], system_df['Date'], pattern)

        # Analyze 8-loss streaks for this system/team (YEARLY reset)
        flags, team_results = analyze_eight_loss_recovery(sys_series, system_df[team], system_df['Date'])

        # Add team/system identifier to results
        for result in team_results:
            result['system'] = system_name
            result['team'] = team
        all_results.extend(team_results)

        # Insert system and flag columns
        tail_idx = list(system_df.columns).index(tail_col) + 1
        system_df.insert(tail_idx, sys_col, sys_series)
        system_df.insert(tail_idx + 1, flag_col, flags)

    # Save system sheet
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        system_df.to_excel(writer, sheet_name=system_name, index=False)

    # Color formatting for this sheet (same as existing systems)
    wb = load_workbook(output_path)
    ws = wb[system_name]
    headers = [cell.value for cell in ws[1]]
    nrows = ws.max_row

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for team in teams:
        team_col = headers.index(team) + 1
        tail_col = headers.index(f"{team} Tail") + 1
        sys_col = headers.index(f"{team} {system_name}") + 1

        for row in range(2, nrows+1):
            ou_val = ws.cell(row=row, column=team_col).value
            tail_val = ws.cell(row=row, column=tail_col).value
            sys_val = ws.cell(row=row, column=sys_col).value

            # O/U coloring
            if ou_val == "Over":
                ws.cell(row=row, column=team_col).fill = green_fill
            elif ou_val == "Under":
                ws.cell(row=row, column=team_col).fill = red_fill
            elif ou_val == "Push":
                ws.cell(row=row, column=team_col).fill = blue_fill
            elif ou_val in ('Skip', '', None):
                ws.cell(row=row, column=team_col).value = "Skip"
                ws.cell(row=row, column=team_col).fill = gray_fill

            # Tail column coloring by outcome
            if tail_val in ("Tail - Over", "Fade - Under"):
                if tail_val == "Tail - Over":
                    if ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
                elif tail_val == "Fade - Under":
                    if ou_val == "Under":
                        ws.cell(row=row, column=tail_col).fill = green_fill
                    elif ou_val == "Over":
                        ws.cell(row=row, column=tail_col).fill = red_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Push":
                ws.cell(row=row, column=tail_col).fill = blue_fill
            elif tail_val == "Skip":
                ws.cell(row=row, column=tail_col).fill = gray_fill

            # System column coloring by O/U result
            if sys_val in ("Tail - Over", "Fade - Under", "Tail - Under", "Fade - Over"):
                if sys_val in ("Tail - Over", "Tail - Under"):
                    # Tailing - win if actual matches what we're tailing
                    if (sys_val == "Tail - Over" and ou_val == "Over") or (sys_val == "Tail - Under" and ou_val == "Under"):
                        ws.cell(row=row, column=sys_col).fill = green_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=sys_col).fill = blue_fill
                    else:
                        ws.cell(row=row, column=sys_col).fill = red_fill
                elif sys_val in ("Fade - Under", "Fade - Over"):
                    # Fading - win if actual is opposite of what we're fading
                    if (sys_val == "Fade - Under" and ou_val == "Under") or (sys_val == "Fade - Over" and ou_val == "Over"):
                        ws.cell(row=row, column=sys_col).fill = green_fill
                    elif ou_val == "Push":
                        ws.cell(row=row, column=sys_col).fill = blue_fill
                    else:
                        ws.cell(row=row, column=sys_col).fill = red_fill
            elif sys_val == "Push":
                ws.cell(row=row, column=sys_col).fill = blue_fill
            elif sys_val == "Skip":
                ws.cell(row=row, column=sys_col).fill = gray_fill

    wb.save(output_path)

# ------------------ STEP 6: CREATE RESULTS SUMMARY ------------------
# Convert accumulated results to DataFrame
results_data = []
for result in all_results:
    results_data.append({
        'System': result['system'],
        'Team': result['team'],
        'Year': result['year'],
        'Start Date': result['start_date'],
        'End Date': result['end_date'],
        'Recovered': result['recovered']
    })

if results_data:
    results_df = pd.DataFrame(results_data)

    # Create summary statistics: count of 8L sequences and how many recovered
    summary = results_df.groupby(['System', 'Team', 'Year']).agg({
        'Recovered': ['count', 'sum']
    }).round(4)

    summary.columns = ['Total_8Loss_Sequences', 'Sequences_Recovered']
    summary = summary.reset_index()
    summary['Recovery_Percentage'] = (summary['Sequences_Recovered'] / summary['Total_8Loss_Sequences'] * 100).round(2)

    # Save both detailed and summary results
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name=results_sheet_name, index=False)
        results_df.to_excel(writer, sheet_name='Detailed Results', index=False)
else:
    # Create empty results sheet if no 8-loss sequences found
    empty_df = pd.DataFrame(columns=['System', 'Team', 'Year', 'Total_8Loss_Sequences', 'Sequences_Recovered', 'Recovery_Percentage'])
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        empty_df.to_excel(writer, sheet_name=results_sheet_name, index=False)

# ------------------ STEP 7: CLEAN UP UNNECESSARY COLUMNS ------------------
wb = load_workbook(output_path)

# Clean up base system sheets (remove {team} Tail and {team} Flag columns)
# ALL COMBINED
base_systems = ["OU TFTF", "OU TFFT", "OU FTTF"]

for sheet_name in base_systems:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        cols_to_delete = []
        for team in teams:
            tail_col_name = f"{team} Tail"
            flag_col_name = f"{team} Flag"
            if tail_col_name in headers:
                cols_to_delete.append(headers.index(tail_col_name) + 1)
            if flag_col_name in headers:
                cols_to_delete.append(headers.index(flag_col_name) + 1)

        for col_idx in sorted(cols_to_delete, reverse=True):
            ws.delete_cols(col_idx)

# Clean up prior system sheets (remove {team} Flag columns only)
# ALL COMBINED
prior_systems = ["TP TFTF", "TP TFFT", "TP FTTF"]

for sheet_name in prior_systems:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        cols_to_delete = []
        for team in teams:
            flag_col_name = f"{team} Flag"
            if flag_col_name in headers:
                cols_to_delete.append(headers.index(flag_col_name) + 1)

        for col_idx in sorted(cols_to_delete, reverse=True):
            ws.delete_cols(col_idx)

wb.save(output_path)

# ------------------ STEP 8: CHECK FOR CURRENT LONG LOSS STREAKS ------------------
def check_current_streaks():
    """
    Scan the most recent results (per sheet/team) and report any active 7+ loss streaks.
    """
    current_streaks = []

    # Check all systems
    all_systems = ['Tails Prior'] + list(system_patterns.keys()) + list(prior_system_patterns.keys())

    for system_name in all_systems:
        sheet_name = tail_sheet_name if system_name == 'Tails Prior' else system_name

        # Read the system sheet
        system_df = pd.read_excel(output_path, sheet_name=sheet_name)
        system_df['Date'] = pd.to_datetime(system_df['Date'])

        for team in teams:
            system_col = f"{team} Tail" if system_name == 'Tails Prior' else f"{team} {system_name}"
            ou_col = team

            team_system_data = system_df[[ou_col, system_col, 'Date']].dropna()
            if len(team_system_data) == 0:
                continue

            # Check current streak by walking backwards from most recent game
            consecutive_losses = 0
            for idx in range(len(team_system_data) - 1, -1, -1):
                system_val = team_system_data.iloc[idx][system_col]
                ou_val = team_system_data.iloc[idx][ou_col]

                # Skip if no valid prediction or push
                if system_val in ('Skip', 'Push', '', None) or pd.isna(system_val) or ou_val in ('Skip', 'Push', '', None) or pd.isna(ou_val):
                    continue

                # Loss conditions
                is_loss = False
                if system_val == "Tail - Over" and ou_val == "Under":
                    is_loss = True
                elif system_val == "Tail - Under" and ou_val == "Over":
                    is_loss = True
                elif system_val == "Fade - Under" and ou_val == "Over":
                    is_loss = True
                elif system_val == "Fade - Over" and ou_val == "Under":
                    is_loss = True

                if is_loss:
                    consecutive_losses += 1
                else:
                    break  # Streak is broken

            # Flag if 7+ losses
            if consecutive_losses >= 7:
                current_streaks.append((team, system_name, consecutive_losses))

    return current_streaks

# Check for current streaks and print results
current_streaks = check_current_streaks()

if current_streaks:
    print("\n" + "="*50)
    print("CURRENT 7+ GAME LOSING STREAKS:")
    print("="*50)
    for team, system, streak_length in current_streaks:
        print(f"{team} is currently on a {streak_length}-game loss streak in the {system} system")
    print("="*50)
else:
    print("\nNo teams are currently on 7+ game losing streaks in any system.")

print(
    f"Finished! Output written to '{output_path}':\n"
    f" - Sheet '{ou_sheet_name}' (O/U results)\n"
    f" - Sheet '{tail_sheet_name}' (Tails Prior system with flags)\n"
    f" - All system pattern sheets with flags\n"
    f" - All prior-based system pattern sheets with flags\n"
    f" - Sheet '{results_sheet_name}' (Recovery analysis summary)\n"
    f" - Sheet 'Detailed Results' (Individual 8-loss sequence details)"
)